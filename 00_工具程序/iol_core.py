"""IOL Tilt Labeler 核心算法与数据层（与 GUI 解耦）。

所有测量口径与旧 Tkinter 版本完全一致：
ImageJ Angle = atan(Height / Width)；结果 = IOL 横线 Angle - A-B 瞳孔平面 Angle。
新增：signed angle、拟合质量分级、自动轴校验、校准测试图。
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
from pathlib import Path
from typing import Any, Iterable

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from platform_utils import resource_path


Point = tuple[float, float]
Circle = tuple[float, float, float]

SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}

MODE_META = {
    "guide": ("1", "C-D 参考线", "可选，多点"),
    "pupil": ("2", "A-B 瞳孔平面", "必须 2 点"),
    "anterior": ("3", "晶体前表面", "至少 3 点，推荐 4 点"),
    "posterior": ("4", "晶体后表面", "至少 3 点，推荐 4-8 点"),
    "manual_axis": ("5", "自动轴校验", "可选 2 点，对照自动轴"),
}

def solve3(matrix: list[list[float]], vector: list[float]) -> list[float]:
    a = [row[:] for row in matrix]
    b = vector[:]
    for i in range(3):
        pivot = max(range(i, 3), key=lambda row: abs(a[row][i]))
        a[i], a[pivot] = a[pivot], a[i]
        b[i], b[pivot] = b[pivot], b[i]
        if abs(a[i][i]) < 1e-12:
            raise ValueError("圆拟合失败：点位可能共线或太接近")
        scale = a[i][i]
        for j in range(i, 3):
            a[i][j] /= scale
        b[i] /= scale
        for row in range(3):
            if row == i:
                continue
            factor = a[row][i]
            for j in range(i, 3):
                a[row][j] -= factor * a[i][j]
            b[row] -= factor * b[i]
    return b


def fit_circle(points: list[Point]) -> Circle:
    if len(points) < 3:
        raise ValueError("拟合圆至少需要 3 个点")

    normal = [[0.0, 0.0, 0.0] for _ in range(3)]
    rhs = [0.0, 0.0, 0.0]
    for x, y in points:
        row = [x, y, 1.0]
        value = -(x * x + y * y)
        for i in range(3):
            rhs[i] += row[i] * value
            for j in range(3):
                normal[i][j] += row[i] * row[j]

    d, e, f = solve3(normal, rhs)
    cx = -d / 2.0
    cy = -e / 2.0
    radius_sq = cx * cx + cy * cy - f
    if radius_sq <= 0:
        raise ValueError("拟合圆半径异常")
    return (cx, cy, math.sqrt(radius_sq))


def circle_intersections(first: Circle, second: Circle) -> list[Point]:
    x0, y0, r0 = first
    x1, y1, r1 = second
    dx = x1 - x0
    dy = y1 - y0
    distance = math.hypot(dx, dy)
    if distance == 0:
        raise ValueError("前后表面拟合圆圆心重合")

    a = (r0 * r0 - r1 * r1 + distance * distance) / (2.0 * distance)
    h_sq = r0 * r0 - a * a
    if h_sq < -1e-6:
        raise ValueError("前后表面拟合圆没有交点，请检查取点")
    h = math.sqrt(max(0.0, h_sq))
    xm = x0 + a * dx / distance
    ym = y0 + a * dy / distance
    rx = -dy / distance * h
    ry = dx / distance * h
    return sorted([(xm + rx, ym + ry), (xm - rx, ym - ry)], key=lambda p: p[0])


def unsigned_imagej_angle(first: Point, second: Point) -> float:
    width = abs(second[0] - first[0])
    height = abs(second[1] - first[1])
    if width == 0:
        return 90.0
    return math.degrees(math.atan(height / width))


def signed_image_angle(first: Point, second: Point) -> float:
    """Signed image angle in degrees. Positive means the line goes downward to the right.

    Image coordinates have y increasing downward, so this matches what the user sees on screen.
    """
    dx = second[0] - first[0]
    dy = second[1] - first[1]
    if dx == 0 and dy == 0:
        return 0.0
    return math.degrees(math.atan2(dy, dx))


def acute_angle_difference(first_angle: float, second_angle: float) -> float:
    diff = (first_angle - second_angle + 90.0) % 180.0 - 90.0
    return diff


def fit_quality(front_rms: float | None, back_rms: float | None) -> tuple[str, str]:
    values = [v for v in (front_rms, back_rms) if v is not None]
    if not values:
        return "manual", "参考轴：无圆拟合误差"
    worst = max(values)
    if worst < 1.0:
        return "good", "✅ 拟合很好（<1 px）"
    if worst < 3.0:
        return "fair", "⚠️ 拟合一般（1-3 px），建议复核点位"
    return "poor", "❌ 拟合偏差较大（>3 px），建议重标"


def imagej_angle(first: Point, second: Point) -> float:
    # Backward-compatible unsigned angle. New code also exports signed angles.
    return unsigned_imagej_angle(first, second)


def width_height(first: Point, second: Point) -> tuple[float, float]:
    return abs(second[0] - first[0]), abs(second[1] - first[1])


def circle_rms_error(points: list[Point], circle: Circle) -> float:
    """拟合质量检查：点到拟合圆的半径误差 RMS，单位 px。"""
    cx, cy, radius = circle
    if not points:
        return 0.0
    errors = [(math.hypot(x - cx, y - cy) - radius) for x, y in points]
    return math.sqrt(sum(e * e for e in errors) / len(errors))


def arc_points(
    circle: Circle,
    x_min: float,
    x_max: float,
    y_target: float,
    y_low: float,
    y_high: float,
    step: int = 2,
) -> list[Point]:
    cx, cy, radius = circle
    pts: list[Point] = []
    for x in range(int(x_min), int(x_max) + 1, step):
        under = radius * radius - (x - cx) * (x - cx)
        if under < 0:
            continue
        root = math.sqrt(under)
        y = min((cy - root, cy + root), key=lambda v: abs(v - y_target))
        if y_low < y < y_high:
            pts.append((float(x), y))
    return pts


def font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    bundled_font = resource_path("fonts", "NotoSansCJKsc-Regular.otf")
    candidates = [bundled_font]
    candidates.extend([
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("/System/Library/Fonts/PingFang.ttc"),
        Path("/System/Library/Fonts/STHeiti Light.ttc"),
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
    ])
    for candidate in candidates:
        try:
            return ImageFont.truetype(str(candidate), size)
        except OSError:
            pass
    return ImageFont.load_default()


def as_points(items: Iterable[Iterable[float]]) -> list[Point]:
    return [(float(x), float(y)) for x, y in items]


def fmt_point(point: Point) -> str:
    return f"({point[0]:.1f},{point[1]:.1f})"


def final_angle_value(result: dict[str, Any]) -> float | None:
    for key in ("final_angle", "abs_difference"):
        if key in result and result[key] not in (None, ""):
            return float(result[key])
    if "difference" in result and result["difference"] not in (None, ""):
        return abs(float(result["difference"]))
    return None


def fmt_number(value: Any, digits: int = 3) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return str(value)


def final_angle_text(result: dict[str, Any]) -> str:
    value = final_angle_value(result)
    return fmt_number(value)


def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    tmp = path.with_name(f".{path.stem}.tmp{path.suffix}")
    tmp.write_text(text, encoding=encoding)
    tmp.replace(path)


def unique_suffix(path: Path) -> str:
    return hashlib.md5(str(path.resolve()).encode("utf-8")).hexdigest()[:8]


def default_label() -> dict[str, Any]:
    return {
        "guide": [],
        "pupil_plane": [],
        "anterior": [],
        "posterior": [],
        "manual_iol_axis": [],
        "result": None,
        "note": "",
    }


def compute_result(label: dict[str, Any]) -> dict[str, Any]:
    anterior = as_points(label.get("anterior", []))
    posterior = as_points(label.get("posterior", []))
    pupil = as_points(label.get("pupil_plane", []))
    manual_axis = as_points(label.get("manual_iol_axis", []))

    if len(pupil) != 2:
        raise ValueError("A-B 瞳孔平面必须正好 2 个点")

    auto_axis: list[Point] | None = None
    front: Circle | None = None
    back: Circle | None = None
    front_rms: float | None = None
    back_rms: float | None = None

    if len(anterior) >= 3 and len(posterior) >= 3:
        front = fit_circle(anterior)
        back = fit_circle(posterior)
        auto_axis = circle_intersections(front, back)
        front_rms = circle_rms_error(anterior, front)
        back_rms = circle_rms_error(posterior, back)
    elif len(manual_axis) != 2:
        if len(anterior) < 3:
            raise ValueError("前表面至少 3 个点；推荐 4 个。或第 5 步自动轴校验点 2 个点。")
        raise ValueError("后表面至少 3 个点；推荐 4-8 个。或第 5 步自动轴校验点 2 个点。")

    if auto_axis is not None:
        iol_l, iol_r = auto_axis
        axis_source = "auto_circle"
    elif len(manual_axis) == 2:
        iol_l, iol_r = sorted([manual_axis[0], manual_axis[1]], key=lambda p: p[0])
        axis_source = "manual_reference"
    else:
        raise ValueError("没有可用的 IOL 轴：请完成前/后表面拟合，或第 5 步自动轴校验点 2 个点")

    iol_angle_signed = signed_image_angle(iol_l, iol_r)
    pupil_angle_signed = signed_image_angle(pupil[0], pupil[1])
    difference = acute_angle_difference(iol_angle_signed, pupil_angle_signed)
    iol_angle = unsigned_imagej_angle(iol_l, iol_r)
    pupil_angle = unsigned_imagej_angle(pupil[0], pupil[1])
    iol_w, iol_h = width_height(iol_l, iol_r)
    pupil_w, pupil_h = width_height(pupil[0], pupil[1])
    quality, quality_note = fit_quality(front_rms, back_rms)

    result: dict[str, Any] = {
        "status": "ok",
        "axis_source": axis_source,
        "iol_axis": [[iol_l[0], iol_l[1]], [iol_r[0], iol_r[1]]],
        "iol_angle": round(iol_angle, 6),
        "pupil_angle": round(pupil_angle, 6),
        "iol_angle_signed": round(iol_angle_signed, 6),
        "pupil_angle_signed": round(pupil_angle_signed, 6),
        "difference": round(difference, 6),
        "final_angle": round(abs(difference), 6),
        "abs_difference": round(abs(difference), 6),
        "iol_width": round(iol_w, 6),
        "iol_height": round(iol_h, 6),
        "pupil_width": round(pupil_w, 6),
        "pupil_height": round(pupil_h, 6),
        "anterior_n": len(anterior),
        "posterior_n": len(posterior),
        "front_fit_rms_px": round(front_rms, 6) if front_rms is not None else "",
        "back_fit_rms_px": round(back_rms, 6) if back_rms is not None else "",
        "fit_quality": quality,
        "fit_quality_note": quality_note,
    }
    if front is not None and back is not None and auto_axis is not None:
        result["front_circle"] = [round(v, 6) for v in front]
        result["back_circle"] = [round(v, 6) for v in back]
        result["auto_iol_axis"] = [[auto_axis[0][0], auto_axis[0][1]], [auto_axis[1][0], auto_axis[1][1]]]
        auto_angle_signed = signed_image_angle(auto_axis[0], auto_axis[1])
        auto_angle = unsigned_imagej_angle(auto_axis[0], auto_axis[1])
        result["auto_iol_angle"] = round(auto_angle, 6)
        result["auto_iol_angle_signed"] = round(auto_angle_signed, 6)
        if len(manual_axis) == 2:
            result["manual_vs_auto_delta"] = round(abs(acute_angle_difference(iol_angle_signed, auto_angle_signed)), 6)
    return result


def save_annotated_image(output_dir: Path, output_stem: str, image_path: Path, label: dict[str, Any], result: dict[str, Any]) -> Path:
    image = Image.open(image_path).convert("RGB")
    draw = ImageDraw.Draw(image)
    main_font = font(24)
    small_font = font(18)
    anterior = as_points(label.get("anterior", []))
    posterior = as_points(label.get("posterior", []))
    pupil = as_points(label.get("pupil_plane", []))
    guide = as_points(label.get("guide", []))
    iol_axis = as_points(result["iol_axis"])
    front = tuple(result["front_circle"]) if result.get("front_circle") else None
    back = tuple(result["back_circle"]) if result.get("back_circle") else None

    # guide
    if len(guide) > 1:
        draw.line(guide, fill=(255, 149, 0), width=2)
    # pupil and iol
    draw.line(pupil, fill=(0, 220, 255), width=2)
    draw.line(iol_axis, fill=(40, 255, 90), width=2)

    # arcs
    if front is not None and back is not None and anterior and posterior:
        y_front = sum(p[1] for p in anterior) / len(anterior)
        y_back = sum(p[1] for p in posterior) / len(posterior)
        x_min = min([p[0] for p in anterior + posterior + iol_axis]) - 40
        x_max = max([p[0] for p in anterior + posterior + iol_axis]) + 40
        front_arc = arc_points(front, x_min, x_max, y_front, y_front - 60, y_front + 60)
        back_arc = arc_points(back, x_min, x_max, y_back, y_back - 60, y_back + 60)
        if len(front_arc) > 1:
            draw.line(front_arc, fill=(245, 220, 0), width=2)
        if len(back_arc) > 1:
            draw.line(back_arc, fill=(255, 149, 0), width=2)

    def marker(point: Point, color: tuple[int, int, int], r: int = 5, text: str | None = None) -> None:
        x, y = point
        draw.ellipse((x - r, y - r, x + r, y + r), fill=(255, 255, 255), outline=color, width=2)
        if text:
            draw.text((x + 8, y - 18), text, fill=color, font=small_font)

    for i, p in enumerate(anterior):
        marker(p, (245, 220, 0), 4, f"F{i+1}")
    for i, p in enumerate(posterior):
        marker(p, (255, 149, 0), 4, f"B{i+1}")
    marker(pupil[0], (0, 220, 255), 6, "A")
    marker(pupil[1], (0, 220, 255), 6, "B")
    marker(iol_axis[0], (255, 40, 40), 6)
    marker(iol_axis[1], (255, 40, 40), 6)

    texts = [
        f"Final Angle = {final_angle_text(result) or 'N/A'}°",
        f"IOL Angle = {fmt_number(result.get('iol_angle'))}°",
        f"A-B Angle = {fmt_number(result.get('pupil_angle'))}°",
        f"Signed IOL - A-B = {fmt_number(result.get('difference'))}°",
        f"Axis = {result.get('axis_source', 'auto_circle')} · {result.get('fit_quality_note', '')}",
    ]
    for idx, text in enumerate(texts):
        x, y = 26, 26 + idx * 32
        bbox = draw.textbbox((x, y), text, font=main_font)
        draw.rectangle((bbox[0] - 6, bbox[1] - 3, bbox[2] + 6, bbox[3] + 3), fill=(0, 0, 0))
        draw.text((x, y), text, fill=(255, 255, 255), font=main_font)

    out = output_dir / "annotated" / f"{output_stem}_annotated.png"
    tmp = out.with_name(f".{out.name}.tmp")
    image.save(tmp, format="PNG")
    tmp.replace(out)


def row_for_image(image: Path, label: dict[str, Any], result: dict[str, Any] | None) -> dict[str, Any]:
    pupil = label.get("pupil_plane", [])
    anterior = label.get("anterior", [])
    posterior = label.get("posterior", [])
    guide = label.get("guide", [])
    if result:
        iol_axis = result.get("iol_axis", [["", ""], ["", ""]])
        return {
            "filename": image.name,
            "source_path": str(image),
            "status": "ok",
            "final_angle": final_angle_text(result),
            "axis_source": result.get("axis_source", "auto_circle"),
            "iol_angle": fmt_number(result.get("iol_angle")),
            "pupil_plane_angle": fmt_number(result.get("pupil_angle")),
            "iol_angle_signed": fmt_number(result.get("iol_angle_signed")),
            "pupil_plane_angle_signed": fmt_number(result.get("pupil_angle_signed")),
            "signed_difference_iol_minus_pupil": fmt_number(result.get("difference")),
            "abs_difference": final_angle_text({"abs_difference": result.get("abs_difference"), "difference": result.get("difference")}),
            "iol_width": fmt_number(result.get("iol_width")),
            "iol_height": fmt_number(result.get("iol_height")),
            "pupil_width": fmt_number(result.get("pupil_width")),
            "pupil_height": fmt_number(result.get("pupil_height")),
            "anterior_n": result.get("anterior_n", len(anterior)),
            "posterior_n": result.get("posterior_n", len(posterior)),
            "front_fit_rms_px": fmt_number(result.get("front_fit_rms_px", 0)),
            "back_fit_rms_px": fmt_number(result.get("back_fit_rms_px", 0)),
            "fit_quality": result.get("fit_quality", ""),
            "fit_quality_note": result.get("fit_quality_note", ""),
            "manual_vs_auto_delta": fmt_number(result.get("manual_vs_auto_delta")),
            "iol_l_x": fmt_number(iol_axis[0][0]),
            "iol_l_y": fmt_number(iol_axis[0][1]),
            "iol_r_x": fmt_number(iol_axis[1][0]),
            "iol_r_y": fmt_number(iol_axis[1][1]),
            "pupil_a_x": f"{pupil[0][0]:.3f}" if len(pupil) == 2 else "",
            "pupil_a_y": f"{pupil[0][1]:.3f}" if len(pupil) == 2 else "",
            "pupil_b_x": f"{pupil[1][0]:.3f}" if len(pupil) == 2 else "",
            "pupil_b_y": f"{pupil[1][1]:.3f}" if len(pupil) == 2 else "",
            "anterior_points": json.dumps(anterior, ensure_ascii=False),
            "posterior_points": json.dumps(posterior, ensure_ascii=False),
            "manual_iol_axis_points": json.dumps(label.get("manual_iol_axis", []), ensure_ascii=False),
            "guide_points": json.dumps(guide, ensure_ascii=False),
            "note": label.get("note", ""),
            "annotated_file": result.get("annotated_file", ""),
        }
    return {
        "filename": image.name,
        "source_path": str(image),
        "status": "unfinished",
        "final_angle": "",
        "axis_source": "",
        "iol_angle": "",
        "pupil_plane_angle": "",
        "iol_angle_signed": "",
        "pupil_plane_angle_signed": "",
        "signed_difference_iol_minus_pupil": "",
        "abs_difference": "",
        "iol_width": "",
        "iol_height": "",
        "pupil_width": "",
        "pupil_height": "",
        "anterior_n": len(anterior),
        "posterior_n": len(posterior),
        "front_fit_rms_px": "",
        "back_fit_rms_px": "",
        "fit_quality": "",
        "fit_quality_note": "",
        "manual_vs_auto_delta": "",
        "iol_l_x": "",
        "iol_l_y": "",
        "iol_r_x": "",
        "iol_r_y": "",
        "pupil_a_x": f"{pupil[0][0]:.3f}" if len(pupil) >= 1 else "",
        "pupil_a_y": f"{pupil[0][1]:.3f}" if len(pupil) >= 1 else "",
        "pupil_b_x": f"{pupil[1][0]:.3f}" if len(pupil) >= 2 else "",
        "pupil_b_y": f"{pupil[1][1]:.3f}" if len(pupil) >= 2 else "",
        "anterior_points": json.dumps(anterior, ensure_ascii=False),
        "posterior_points": json.dumps(posterior, ensure_ascii=False),
        "manual_iol_axis_points": json.dumps(label.get("manual_iol_axis", []), ensure_ascii=False),
        "guide_points": json.dumps(guide, ensure_ascii=False),
        "note": label.get("note", ""),
        "annotated_file": "",
    }


def write_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "IOL Tilt Results"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
    for col_idx, header in enumerate(headers, start=1):
        width = min(45, max(12, len(header) + 2))
        if header in {"filename", "source_path", "annotated_file", "anterior_points", "posterior_points", "guide_points"}:
            width = 34
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    tmp = path.with_name(f".{path.stem}.tmp{path.suffix}")
    wb.save(tmp)
    tmp.replace(path)


def generate_calibration_images(output_dir: Path | None = None) -> Path:
    if output_dir is None:
        out_dir = Path.home() / "IOL_Tilt_Labeler" / "IOL校准测试图"
    else:
        out_dir = output_dir.parent / "IOL校准测试图"
    out_dir.mkdir(parents=True, exist_ok=True)
    cases = [
        ("pupil0_iol5", 0.0, 5.0),
        ("pupil0_iol10", 0.0, 10.0),
        ("pupil0_iol-10", 0.0, -10.0),
        ("pupil3_iol12", 3.0, 12.0),
        ("pupil-4_iol8", -4.0, 8.0),
        ("pupil7_iol-6", 7.0, -6.0),
    ]
    rows = [["filename", "pupil_angle_signed", "iol_angle_signed", "expected_signed_difference", "expected_abs_difference", "note"]]

    def line_points(cx: float, cy: float, length: float, angle_deg: float) -> tuple[Point, Point]:
        rad = math.radians(angle_deg)
        dx = math.cos(rad) * length / 2.0
        dy = math.sin(rad) * length / 2.0
        return (cx - dx, cy - dy), (cx + dx, cy + dy)

    for name, pupil_angle, iol_angle in cases:
        image = Image.new("RGB", (1200, 760), (12, 15, 24))
        draw = ImageDraw.Draw(image)
        main = font(26)
        small = font(18)
        pupil = line_points(600, 385, 760, pupil_angle)
        iol = line_points(600, 360, 560, iol_angle)
        # Decorative lens surfaces only for visual practice; the purple line is the exact answer.
        for y, color in [(330, (245, 220, 0)), (390, (255, 149, 0))]:
            pts = []
            for x in range(320, 881, 8):
                curve = 0.00055 * (x - 600) * (x - 600)
                pts.append((x, y + curve))
            draw.line(pts, fill=color, width=3)
        draw.line(pupil, fill=(0, 220, 255), width=4)
        draw.line(iol, fill=(191, 90, 242), width=5)
        for pnt, txt, color in [(pupil[0], "A", (0,220,255)), (pupil[1], "B", (0,220,255)), (iol[0], "M1", (191,90,242)), (iol[1], "M2", (191,90,242))]:
            x, y = pnt
            draw.ellipse((x-8, y-8, x+8, y+8), fill=(255,255,255), outline=color, width=3)
            draw.text((x+12, y-22), txt, fill=color, font=small)
        diff = acute_angle_difference(iol_angle, pupil_angle)
        texts = [
            "IOL Tilt 校准测试图",
            f"A-B signed = {pupil_angle:.3f}°",
            f"IOL signed = {iol_angle:.3f}°",
            f"Expected IOL - A-B = {diff:.3f}°  |abs| = {abs(diff):.3f}°",
            "提示：第 2 步点 A/B；第 5 步点紫色 M1/M2，可先验证 signed angle。",
        ]
        for idx, text in enumerate(texts):
            draw.text((36, 32 + idx * 34), text, fill=(255,255,255), font=main if idx == 0 else small)
        filename = f"{name}.png"
        image.save(out_dir / filename)
        rows.append([filename, f"{pupil_angle:.6f}", f"{iol_angle:.6f}", f"{diff:.6f}", f"{abs(diff):.6f}", "紫色 M1/M2 为标准 IOL 轴；黄/橙弧仅作视觉练习"])
    csv_text = "\n".join(",".join(row) for row in rows) + "\n"
    atomic_write_text(out_dir / "00_标准答案.csv", csv_text, encoding="utf-8-sig")
    return out_dir
